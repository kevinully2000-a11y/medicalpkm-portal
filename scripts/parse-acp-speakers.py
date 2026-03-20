#!/usr/bin/env python3
"""
Parse ACP IMM 2026 Sessions spreadsheet → enriched CSV for batch generation.
- Extracts unique speakers with affiliations and headshot URLs
- Looks up NPIs via NPPES Registry API
- Assigns priority based on leadership positions
- Outputs batch-ready CSV
"""

import json
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

XLSX_PATH = Path.home() / "Downloads" / "ACP_IMM_2026_Sessions.xlsx"
OUTPUT_CSV = Path(__file__).parent.parent / "data" / "acp-imm-2026-speakers.csv"

# Leadership keywords and their point values for priority scoring
LEADERSHIP_SCORES = {
    'chair': 3, 'chairman': 3, 'chairwoman': 3, 'chairperson': 3,
    'chief': 3, 'dean': 3, 'vice dean': 3, 'associate dean': 3,
    'president': 3, 'vice president': 3,
    'director': 2, 'co-director': 2,
    'professor of medicine': 1,  # full professor (not assoc/asst)
    'macp': 2,  # Master of ACP - highest ACP honor
}


def parse_name(full_name: str) -> dict:
    """Parse 'Julia Arnsten, MD, FACP' → {'full': ..., 'first': ..., 'last': ..., 'credentials': ...}"""
    # Split on comma — credentials come after the first comma that's followed by a known credential
    stripped = full_name.strip()
    name_part = stripped
    creds = ''

    # Look for ", MD" or ", DO" etc. as a word boundary (not inside a name like Eduardo)
    cred_match = re.search(
        r',\s*((?:MD|M\.D\.|DO|D\.O\.|PhD|MPH|MHA|MHPE|MSEd|MSc|MS|MSCE|MGIN|MBBS|JD|LCSW|APHSW-C|COTA/?L?|LPC|RN|NP|PA|FACP|MACP|FACS|FACEP|FRCP|Member|Jr\.|Sr\.)(?:\b|,).*)',
        stripped
    )
    if cred_match:
        cred_start = cred_match.start()
        name_part = stripped[:cred_start].strip()
        creds = cred_match.group(1).strip()
    else:
        # No credentials found — check if there's a comma at all
        if ',' in stripped:
            # Could be "Last, First" format — keep as-is
            name_part = stripped

    parts = name_part.split()
    first = parts[0] if parts else ''
    last = parts[-1] if len(parts) > 1 else parts[0] if parts else ''

    return {
        'full': full_name.strip(),
        'clean_name': name_part,
        'first': first,
        'last': last,
        'credentials': creds.strip().lstrip(',').strip(),
    }


def extract_institution(affiliation: str) -> str:
    """Extract institution name from full affiliation text."""
    # Remove the 'Has no relationships...' suffix
    clean = affiliation.split('Has no relationships')[0].strip().rstrip(',').rstrip('.')

    # Try to find institution by keyword matching
    parts = [p.strip() for p in clean.split(',')]
    institution_keywords = ['University', 'Hospital', 'Medical Center', 'College',
                           'Institute', 'Clinic', 'School', 'Health', 'VA ']
    for p in reversed(parts):  # institutions tend to be later in the string
        if any(kw in p for kw in institution_keywords):
            # Include the city/state that follows if present
            idx = parts.index(p)
            result = p
            # Check if next parts are city/state
            if idx + 1 < len(parts) and len(parts[idx + 1]) <= 30:
                result += ', ' + parts[idx + 1]
            return result

    # Fallback: return everything after the first comma (skip the role/title)
    if len(parts) > 1:
        return ', '.join(parts[1:4])
    return clean[:100]


def extract_state(affiliation: str) -> str:
    """Extract 2-letter state code from affiliation."""
    match = re.search(r',\s*([A-Z]{2})(?:\s*\.|,|\s+Has\s|\s*$)', affiliation)
    return match.group(1) if match else ''


def extract_specialty(affiliation: str) -> str:
    """Extract specialty/division from affiliation text."""
    # Look for "Division of X" or "Department of X"
    match = re.search(r'Division of ([^,]+)', affiliation)
    if match:
        return match.group(1).strip()
    match = re.search(r'Department of ([^,]+)', affiliation)
    if match:
        return match.group(1).strip()
    return ''


def score_priority(affiliation: str, credentials: str, session_count: int) -> tuple:
    """Score a speaker's priority based on leadership. Returns (score, priority_level)."""
    score = 0
    aff_lower = affiliation.lower()

    for keyword, points in LEADERSHIP_SCORES.items():
        if keyword in aff_lower:
            # 'professor of medicine' should only match full professor, not associate/assistant
            if keyword == 'professor of medicine':
                if 'associate professor' in aff_lower or 'assistant professor' in aff_lower:
                    continue
            score += points

    # Bonus for MACP in credentials
    if 'MACP' in credentials:
        score += 2

    # Bonus for multi-session speakers
    if session_count >= 2:
        score += 1

    if score >= 3:
        return score, 'high'
    elif score >= 1:
        return score, 'medium'
    return score, 'low'


def lookup_npi(first_name: str, last_name: str, state: str = '', retries: int = 2) -> str:
    """Look up NPI from NPPES Registry. Returns NPI string or empty."""
    params = {
        'version': '2.1',
        'first_name': first_name,
        'last_name': last_name,
        'enumeration_type': 'NPI-1',
        'limit': '5',
    }
    if state:
        params['state'] = state

    url = f'https://npiregistry.cms.hhs.gov/api/?{urllib.parse.urlencode(params)}'

    for attempt in range(retries + 1):
        try:
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read())
                if data.get('result_count', 0) == 0:
                    return ''
                results = data.get('results', [])
                if len(results) == 1:
                    return str(results[0]['number'])
                # Multiple results — try to disambiguate by state
                if state:
                    for r in results:
                        addrs = r.get('addresses', [])
                        if any(a.get('state') == state for a in addrs):
                            return str(r['number'])
                # If still ambiguous, check if the first result is an MD
                for r in results:
                    cred = r.get('basic', {}).get('credential', '')
                    if 'M.D.' in cred or 'MD' in cred or 'D.O.' in cred:
                        return str(r['number'])
                return str(results[0]['number']) if results else ''
        except Exception as e:
            if attempt < retries:
                time.sleep(1)
                continue
            print(f'  NPI lookup failed for {first_name} {last_name}: {e}', file=sys.stderr)
            return ''


def escape_csv_field(value: str) -> str:
    """Escape a CSV field — quote if it contains commas, quotes, or newlines."""
    if ',' in value or '"' in value or '\n' in value:
        return '"' + value.replace('"', '""') + '"'
    return value


def main():
    try:
        import openpyxl
    except ImportError:
        # Try pandas as fallback
        try:
            import pandas as pd
        except ImportError:
            print('ERROR: Need openpyxl or pandas. Run: pip install openpyxl', file=sys.stderr)
            sys.exit(1)

        df = pd.read_excel(str(XLSX_PATH))
        rows_data = []
        for _, row in df.iterrows():
            rows_data.append({col: row[col] for col in df.columns})
    else:
        wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
        ws = wb.active
        headers = [str(cell.value or '') for cell in ws[1]]
        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows_data.append({headers[i]: (row[i] if row[i] is not None else '') for i in range(len(headers))})

    # Extract unique speakers
    speakers = {}
    for row in rows_data:
        for i in range(1, 6):
            name_col = f'Speaker {i}'
            aff_col = f'Affiliation {i}'
            url_col = f'Speaker {i} Headshot URL'

            name = str(row.get(name_col, '') or '').strip()
            if not name:
                continue

            if name not in speakers:
                aff = str(row.get(aff_col, '') or '')
                url = str(row.get(url_col, '') or '')
                speakers[name] = {
                    'affiliation': aff,
                    'headshot_url': url,
                    'sessions': [],
                }
            session_title = str(row.get('Session Title', '') or '')
            if session_title:
                speakers[name]['sessions'].append(session_title)

    print(f'Found {len(speakers)} unique speakers')

    # Process each speaker
    results = []
    for idx, (raw_name, data) in enumerate(sorted(speakers.items())):
        parsed = parse_name(raw_name)
        aff = data['affiliation']
        institution = extract_institution(aff)
        state = extract_state(aff)
        specialty = extract_specialty(aff)
        session_count = len(data['sessions'])
        score, priority = score_priority(aff, parsed['credentials'], session_count)

        # NPI lookup with rate limiting
        print(f'  [{idx+1}/{len(speakers)}] {parsed["clean_name"]} ({state}) ...', end=' ', flush=True)
        npi = lookup_npi(parsed['first'], parsed['last'], state)
        print(f'NPI: {npi or "NOT FOUND"} | Priority: {priority} (score={score})')
        time.sleep(0.4)  # Rate limiting

        # Build additional context from session titles
        context_parts = [f'ACP Internal Medicine Meeting 2026 speaker']
        for s in data['sessions'][:3]:
            context_parts.append(f'Session: {s}')

        results.append({
            'kolName': parsed['clean_name'],
            'institution': institution,
            'specialty': specialty,
            'npi': npi,
            'conference': 'ACP.26',
            'priority': priority,
            'headshotUrl': data['headshot_url'],
            'additionalContext': '; '.join(context_parts),
        })

    # Write CSV
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CSV, 'w') as f:
        headers = ['kolName', 'institution', 'specialty', 'npi', 'conference', 'priority', 'headshotUrl', 'additionalContext']
        f.write(','.join(headers) + '\n')
        for r in results:
            row = [escape_csv_field(str(r.get(h, ''))) for h in headers]
            f.write(','.join(row) + '\n')

    print(f'\nWrote {len(results)} rows to {OUTPUT_CSV}')

    # Summary
    priorities = {'high': 0, 'medium': 0, 'low': 0}
    npi_found = 0
    for r in results:
        priorities[r['priority']] += 1
        if r['npi']:
            npi_found += 1

    print(f'\nPriority distribution: HIGH={priorities["high"]}, MEDIUM={priorities["medium"]}, LOW={priorities["low"]}')
    print(f'NPI match rate: {npi_found}/{len(results)} ({npi_found/len(results)*100:.0f}%)')


if __name__ == '__main__':
    main()
